"""Extrae los datos de PorraMundial2026.xlsx a archivos JSON en /data."""
import json, warnings, unicodedata, re
from pathlib import Path
import openpyxl

warnings.filterwarnings("ignore")

SRC = Path(r"C:\Users\vicen\Downloads\PorraMundial2026.xlsx")
OUT = Path(__file__).resolve().parent.parent / "data"
OUT.mkdir(exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True)


def slug(text):
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


# ---------- Equipos y grupos ----------
ws = wb["Grupos"]
teams = []
# 3 bloques de columnas: B-D (grupos A-D), F-H (E-H), J-L (I-L)
blocks = [(2, 3, 4), (6, 7, 8), (10, 11, 12)]
for col_g, col_t, col_c in blocks:
    current_group = None
    for r in range(3, ws.max_row + 1):
        g = ws.cell(row=r, column=col_g).value
        t = ws.cell(row=r, column=col_t).value
        c = ws.cell(row=r, column=col_c).value
        if g:
            current_group = str(g).strip()
        if t and current_group:
            teams.append({
                "id": slug(t),
                "name": str(t).strip(),
                "group": current_group,
                "confederation": str(c).strip() if c else "",
            })

groups = []
for letter in sorted({t["group"] for t in teams}):
    groups.append({
        "id": letter,
        "name": f"Grupo {letter}",
        "teams": [t["id"] for t in teams if t["group"] == letter],
    })

# ---------- Sedes ----------
ws = wb["Sedes"]
venues = []
for r in range(3, ws.max_row + 1):
    name = ws.cell(row=r, column=2).value
    if not name:
        continue
    venues.append({
        "id": slug(ws.cell(row=r, column=3).value),
        "stadium": str(name).strip(),
        "officialName": str(ws.cell(row=r, column=3).value).strip(),
        "city": str(ws.cell(row=r, column=4).value).strip(),
        "country": str(ws.cell(row=r, column=5).value).strip(),
        "capacity": int(ws.cell(row=r, column=6).value or 0),
    })

# ---------- Partidos de fase de grupos ----------
ws = wb["Fase de Grupos"]
matches = []
current_group = None
counters = {}
for r in range(6, ws.max_row + 1):
    g = ws.cell(row=r, column=2).value
    date = ws.cell(row=r, column=3).value
    home = ws.cell(row=r, column=7).value
    away = ws.cell(row=r, column=10).value
    if g:
        current_group = str(g).replace("Grupo", "").strip()
    if not (home and away):
        continue
    time = ws.cell(row=r, column=4).value
    counters[current_group] = counters.get(current_group, 0) + 1
    n = counters[current_group]
    matches.append({
        "id": f"{current_group}-{n}",
        "group": current_group,
        "date": date.strftime("%Y-%m-%d") if hasattr(date, "strftime") else str(date),
        "time": time.strftime("%H:%M") if hasattr(time, "strftime") else str(time or ""),
        "city": str(ws.cell(row=r, column=5).value or "").strip(),
        "venue": str(ws.cell(row=r, column=6).value or "").strip(),
        "home": slug(home),
        "away": slug(away),
        "homeName": str(home).strip(),
        "awayName": str(away).strip(),
    })

# ---------- Cuadro final (bracket) ----------
# Estructura parseada de la hoja "Fase Final" (slots tipo "1ºE", "3ºABCDF").
bracket = {
    "r32": [
        {"id": "D1", "slots": ["1ºE", "3ºABCDF"], "date": "2026-06-29", "city": "Boston"},
        {"id": "D2", "slots": ["1ºI", "3ºCDFGH"], "date": "2026-06-30", "city": "Nueva York"},
        {"id": "D3", "slots": ["2ºA", "2ºB"], "date": "2026-06-28", "city": "Los Ángeles"},
        {"id": "D4", "slots": ["1ºF", "2ºC"], "date": "2026-06-29", "city": "Monterrey"},
        {"id": "D5", "slots": ["2ºK", "2ºL"], "date": "2026-07-02", "city": "Toronto"},
        {"id": "D6", "slots": ["1ºH", "2ºJ"], "date": "2026-07-02", "city": "Los Ángeles"},
        {"id": "D7", "slots": ["1ºD", "3ºBEFIJ"], "date": "2026-07-01", "city": "San Francisco"},
        {"id": "D8", "slots": ["1ºG", "3ºAEHIJ"], "date": "2026-07-01", "city": "Seattle"},
        {"id": "D9", "slots": ["1ºC", "2ºF"], "date": "2026-06-29", "city": "Houston"},
        {"id": "D10", "slots": ["2ºE", "2ºI"], "date": "2026-06-30", "city": "Dallas"},
        {"id": "D11", "slots": ["1ºA", "3ºCEFHI"], "date": "2026-06-30", "city": "Ciudad de México"},
        {"id": "D12", "slots": ["1ºL", "3ºEHIJK"], "date": "2026-07-01", "city": "Atlanta"},
        {"id": "D13", "slots": ["1ºJ", "2ºH"], "date": "2026-07-03", "city": "Miami"},
        {"id": "D14", "slots": ["2ºD", "2ºG"], "date": "2026-07-03", "city": "Dallas"},
        {"id": "D15", "slots": ["1ºB", "3ºEFGIJ"], "date": "2026-07-02", "city": "Vancouver"},
        {"id": "D16", "slots": ["1ºK", "3ºDEIJL"], "date": "2026-07-03", "city": "Kansas City"},
    ],
    "r16": [
        {"id": "O1", "from": ["D1", "D2"], "date": "2026-07-04", "city": "Philadelphia"},
        {"id": "O2", "from": ["D3", "D4"], "date": "2026-07-04", "city": "Houston"},
        {"id": "O3", "from": ["D5", "D6"], "date": "2026-07-06", "city": "Dallas"},
        {"id": "O4", "from": ["D7", "D8"], "date": "2026-07-06", "city": "Seattle"},
        {"id": "O5", "from": ["D9", "D10"], "date": "2026-07-05", "city": "Nueva York"},
        {"id": "O6", "from": ["D11", "D12"], "date": "2026-07-05", "city": "Ciudad de México"},
        {"id": "O7", "from": ["D13", "D14"], "date": "2026-07-07", "city": "Atlanta"},
        {"id": "O8", "from": ["D15", "D16"], "date": "2026-07-07", "city": "Vancouver"},
    ],
    "qf": [
        {"id": "Q1", "from": ["O1", "O2"], "date": "2026-07-09", "city": "Boston"},
        {"id": "Q2", "from": ["O3", "O4"], "date": "2026-07-10", "city": "Los Ángeles"},
        {"id": "Q3", "from": ["O5", "O6"], "date": "2026-07-11", "city": "Miami"},
        {"id": "Q4", "from": ["O7", "O8"], "date": "2026-07-11", "city": "Kansas City"},
    ],
    "sf": [
        {"id": "S1", "from": ["Q1", "Q2"], "date": "2026-07-14", "city": "Dallas"},
        {"id": "S2", "from": ["Q3", "Q4"], "date": "2026-07-15", "city": "Atlanta"},
    ],
    "third": {"id": "T1", "from": ["S1", "S2"], "date": "2026-07-18", "city": "Miami"},
    "final": {"id": "F1", "from": ["S1", "S2"], "date": "2026-07-19", "city": "Nueva York"},
}

# ---------- Preguntas ----------
ws = wb["PREGUNTAS"]
questions = []
category = "Generales"
BOOL_HINTS = ("¿habrá", "¿ganará", "¿puntuará", "¿marcará", "¿jugará", "¿conseguirá",
              "¿pasará", "¿parará", "¿logrará", "¿se producirá")
for r in range(4, 50):
    text = ws.cell(row=r, column=2).value
    points = ws.cell(row=r, column=4).value
    if not text:
        continue
    text = str(text).strip()
    if points in (None, ""):  # fila de categoría
        category = text
        continue
    low = text.lower()
    if low.startswith(BOOL_HINTS):
        qtype = "boolean"
    elif "qué selección" in low or "que seleccion" in low:
        qtype = "team"
    else:
        qtype = "text"
    questions.append({
        "id": f"q{r}",
        "category": category,
        "text": text,
        "points": int(points),
        "type": qtype,
    })

# ---------- Escritura ----------
def write(name, data):
    path = OUT / name
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  {name}: {len(data) if isinstance(data, list) else 'ok'}")

print("Datos extraídos a /data:")
write("teams.json", teams)
write("groups.json", groups)
write("venues.json", venues)
write("matches.json", matches)
write("bracket.json", bracket)
write("questions.json", questions)
print(f"\nTotal: {len(teams)} equipos, {len(groups)} grupos, {len(venues)} sedes, "
      f"{len(matches)} partidos, {len(questions)} preguntas.")
